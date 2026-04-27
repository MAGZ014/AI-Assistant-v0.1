# tools/files/excel_files.py
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side
)
from openpyxl.utils import get_column_letter
from config import DOCS_FOLDER


def _build_path(filename: str) -> tuple:
    """Asegura extensión .xlsx y regresa (ruta, filename)."""
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    return os.path.join(DOCS_FOLDER, filename), filename


def _auto_adjust_columns(ws):
    """Ajusta el ancho de columnas automáticamente según el contenido."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def _style_header_row(ws, num_cols: int):
    """Aplica estilo al encabezado (fila 1)."""
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _style_data_rows(ws, num_rows: int, num_cols: int):
    """Aplica estilo alternado a las filas de datos."""
    fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = fill_even
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border


# ─────────────────────────────────────────
#  CREAR
# ─────────────────────────────────────────

def create_xlsx(filename: str, headers: list, rows: list = None, sheet_name: str = "Hoja1") -> str:
    """
    Crea un Excel con encabezados y datos opcionales.

    headers = ["Nombre", "Edad", "Ciudad"]
    rows    = [["Ana", 25, "Monterrey"], ["Luis", 30, "CDMX"]]
    """
    path, filename = _build_path(filename)
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Encabezados
        ws.append(headers)
        _style_header_row(ws, len(headers))

        # Datos
        if rows:
            for row in rows:
                ws.append(row)
            _style_data_rows(ws, len(rows), len(headers))

        # Freeze header row
        ws.freeze_panes = "A2"

        _auto_adjust_columns(ws)
        wb.save(path)
        return f"Excel '{filename}' creado con {len(rows) if rows else 0} filas."
    except Exception as e:
        return f"Error creando '{filename}': {str(e)}"


def create_xlsx_from_dict(filename: str, data: list, sheet_name: str = "Hoja1") -> str:
    """
    Crea un Excel desde una lista de diccionarios.
    Las keys del primer dict se usan como encabezados.

    data = [
        {"Nombre": "Ana", "Edad": 25, "Ciudad": "Monterrey"},
        {"Nombre": "Luis", "Edad": 30, "Ciudad": "CDMX"},
    ]
    """
    if not data:
        return "No hay datos para crear el Excel."

    headers = list(data[0].keys())
    rows = [[item.get(h, "") for h in headers] for item in data]
    return create_xlsx(filename, headers, rows, sheet_name)


def create_xlsx_multi_sheet(filename: str, sheets: list) -> str:
    """
    Crea un Excel con múltiples hojas.

    sheets = [
        {"name": "Enero", "headers": ["Producto", "Ventas"], "rows": [["Laptop", 5000]]},
        {"name": "Febrero", "headers": ["Producto", "Ventas"], "rows": [["Mouse", 200]]},
    ]
    """
    path, filename = _build_path(filename)
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Elimina hoja default

        for sheet in sheets:
            ws = wb.create_sheet(title=sheet.get("name", "Hoja"))
            headers = sheet.get("headers", [])
            rows    = sheet.get("rows", [])

            ws.append(headers)
            _style_header_row(ws, len(headers))

            for row in rows:
                ws.append(row)

            if rows:
                _style_data_rows(ws, len(rows), len(headers))

            ws.freeze_panes = "A2"
            _auto_adjust_columns(ws)

        wb.save(path)
        return f"Excel '{filename}' creado con {len(sheets)} hojas."
    except Exception as e:
        return f"Error creando '{filename}': {str(e)}"


# ─────────────────────────────────────────
#  LEER
# ─────────────────────────────────────────

def read_xlsx(filename: str, sheet_name: str = None, max_rows: int = 50) -> str:
    """
    Lee un Excel y regresa su contenido como texto.
    Si no se especifica hoja, lee la primera.
    """
    path, filename = _build_path(filename)

    if not os.path.exists(path):
        return f"No encontré '{filename}'."

    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Seleccionar hoja
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        lines = [f"Hoja: {ws.title}"]
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue
            lines.append(" | ".join([str(v) if v is not None else "" for v in row]))
            row_count += 1
            if row_count >= max_rows:
                lines.append(f"... [truncado en {max_rows} filas]")
                break

        # Info de hojas disponibles
        if len(wb.sheetnames) > 1:
            lines.append(f"\nHojas disponibles: {', '.join(wb.sheetnames)}")

        return "\n".join(lines)
    except Exception as e:
        return f"Error leyendo '{filename}': {str(e)}"


def list_sheets(filename: str) -> str:
    """Lista las hojas disponibles en un Excel."""
    path, filename = _build_path(filename)

    if not os.path.exists(path):
        return f"No encontré '{filename}'."

    try:
        wb = openpyxl.load_workbook(path)
        return f"Hojas en '{filename}': {', '.join(wb.sheetnames)}"
    except Exception as e:
        return f"Error: {str(e)}"


# ─────────────────────────────────────────
#  MODIFICAR
# ─────────────────────────────────────────

def append_rows_xlsx(filename: str, rows: list, sheet_name: str = None) -> str:
    """Agrega filas al final de un Excel existente."""
    path, filename = _build_path(filename)

    if not os.path.exists(path):
        return f"No encontré '{filename}'."

    try:
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        for row in rows:
            ws.append(row)

        _auto_adjust_columns(ws)
        wb.save(path)
        return f"{len(rows)} fila(s) agregadas a '{filename}'."
    except Exception as e:
        return f"Error modificando '{filename}': {str(e)}"


def add_sheet_xlsx(filename: str, sheet_name: str, headers: list, rows: list = None) -> str:
    """Agrega una nueva hoja a un Excel existente."""
    path, filename = _build_path(filename)

    if not os.path.exists(path):
        return f"No encontré '{filename}'."

    try:
        wb = openpyxl.load_workbook(path)

        if sheet_name in wb.sheetnames:
            return f"Ya existe una hoja llamada '{sheet_name}' en '{filename}'."

        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        _style_header_row(ws, len(headers))

        if rows:
            for row in rows:
                ws.append(row)
            _style_data_rows(ws, len(rows), len(headers))

        ws.freeze_panes = "A2"
        _auto_adjust_columns(ws)
        wb.save(path)
        return f"Hoja '{sheet_name}' agregada a '{filename}'."
    except Exception as e:
        return f"Error: {str(e)}"